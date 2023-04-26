import sys
import os
import openpyxl
import folium
from PyQt5.QtCore import QUrl
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QTableWidget,
    QTableWidgetItem,
    QFileDialog,
    QTabWidget,
    QAction,
    QMessageBox,
    QInputDialog
)
from PyQt5.QtWebEngineWidgets import QWebEngineView
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # Creazione della finestra principale
        self.setWindowTitle("Visualizzatore di punti geografici")
        self.setGeometry(100, 100, 800, 600)

        # Creazione del widget per visualizzare la mappa
        self.mapTabs = QTabWidget(self)
        self.setCentralWidget(self.mapTabs)

        # Creazione della barra dei menu
        menubar = self.menuBar()

        # Creazione del menu File e delle relative azioni
        fileMenu = menubar.addMenu('File')

        openFileAction = QAction('Apri', self)
        openFileAction.setShortcut('Ctrl+O')
        openFileAction.triggered.connect(self.open_file)
        fileMenu.addAction(openFileAction)

        printMapAction = QAction('Stampa la mappa', self)
        printMapAction.setShortcut('Ctrl+P')
        printMapAction.triggered.connect(self.print_map)
        fileMenu.addAction(printMapAction)

        # Creazione del menu Visualizza e delle relative azioni
        viewMenu = menubar.addMenu('Visualizza')

        showCoordinatesAction = QAction('Mostra le coordinate', self)
        showCoordinatesAction.setShortcut('Ctrl+C')
        showCoordinatesAction.triggered.connect(self.show_coordinates)
        viewMenu.addAction(showCoordinatesAction)

        showOnMapAction = QAction('Mostra sulla mappa', self)
        showOnMapAction.setShortcut('Ctrl+M')
        showOnMapAction.triggered.connect(self.show_on_map)
        viewMenu.addAction(showOnMapAction)

        # Creazione della barra di stato
        self.statusBar()


    def load_data(self):
        # Apertura della finestra di dialogo per selezionare il file Excel
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filename, _ = QFileDialog.getOpenFileName(
            self, "Seleziona un file Excel", "", "Excel Files (*.xlsx);;", options=options)
        if not filename:
            return

        # Lettura del file Excel
        workbook = openpyxl.load_workbook(filename, read_only=True)

        # Creazione del widget per visualizzare il file Excel
        excelTabs = QTabWidget(self)
        self.mapTabs.addTab(excelTabs, "Excel")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            tableWidget = QTableWidget()
            tableWidget.setColumnCount(worksheet.max_column)
            #tableWidget.setRowCount(worksheet.max_row)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                tableWidget.insertRow(tableWidget.rowCount())
                for column, cell_value in enumerate(row):
                    item = QTableWidgetItem(str(cell_value))
                    tableWidget.setItem(
                        tableWidget.rowCount() - 1, column, item)
            tableWidget.setHorizontalHeaderLabels(
                tuple(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0])
            excelTabs.addTab(tableWidget, sheet_name)

        # Calcolo del bounding box per tutti i siti
        min_lat, max_lat, min_lon, max_lon = None, None, None, None
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                lat = row[1]
                lon = row[2]
                if min_lat is None or lat < min_lat:
                    min_lat = lat
                if max_lat is None or lat > max_lat:
                    max_lat = lat
                if min_lon is None or lon < min_lon:
                    min_lon = lon
                if max_lon is None or lon > max_lon:
                    max_lon = lon

        # Creazione della mappa per ogni foglio di lavoro
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Creazione del layer per i punti
            layer = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                name = row[0]
                lat = row[1]
                lon = row[2]
                layer.append({
                    "name": name,
                    "lat": lat,
                    "lon": lon
                })

            # Creazione della mappa per il foglio di lavoro
            self.map_html = folium.Map(location=[(min_lat + max_lat) / 2, (min_lon + max_lon) / 2], zoom_start=10)
            self.map_html.fit_bounds([[min_lat, min_lon], [max_lat, max_lon]])
            # Inserimento del markup nella mappa


            for point in layer:
                folium.Marker(location=[float(point['lat']), float(point['lon'])], popup=point['name']).add_to(
                    self.map_html)

            # Inserimento del markup nella mappa
            map_path = os.path.abspath("map.html")
            self.map_html.save(map_path)
            print("Percorso del file map.html:", map_path)

            # Creazione del widget per la mappa
            self.mapWidget = QWebEngineView()
            self.mapWidget.load(QUrl.fromLocalFile(map_path))
            self.mapTabs.addTab(self.mapWidget, sheet_name)

            # Aggiunta dei marker per i punti
            for point in layer:
                name = point['name']
                lat = point['lat']
                lon = point['lon']
                self.mapWidget.page().runJavaScript(f'''
                    L.marker([{lat}, {lon}]).addTo(map).bindPopup("{name}");
                ''')

            self.mapWidget.page().runJavaScript(f'''
                map.fitBounds([
                    [{min_lat}, {min_lon}],
                    [{max_lat}, {max_lon}]
                ]);
            ''')

            # Aggiunta della mappa alla tab
            self.mapTabs.addTab(self.mapWidget, sheet_name)

    def open_file(self):
        self.load_data()

    def print_map(self):
        # Salva i contenuti della vista della mappa come un'immagine
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filename, _ = QFileDialog.getSaveFileName(
            self, "Salva la mappa", "mappa", "JPEG Files (*.jpg);;", options=options)
        if not filename:
            return
        pixmap = self.mapTabs.currentWidget().grab()
        pixmap.save(filename, "jpg", quality=100)

        # Chiedi all'utente la risoluzione dell'immagine in dpi
        dpi, ok = QInputDialog.getInt(
            self, "Risoluzione", "Inserisci la risoluzione dell'immagine in dpi", 300, 1, 1000)
        if not ok:
            return

        # Crea un messaggio di conferma per l'utente
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Mappa stampata")
        msgBox.setText(f"La mappa è stata stampata con successo a {dpi} dpi.")
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec()

    def show_coordinates(self):
        # Recupera la tabella selezionata e la riga selezionata
        currentTab = self.mapTabs.currentIndex()
        currentTable = self.mapTabs.widget(currentTab).currentWidget()
        selectedRows = currentTable.selectionModel().selectedRows()
        if len(selectedRows) == 0:
            return

        # Recupera le coordinate del punto selezionato
        selectedRow = selectedRows[0].row()
        latitude = float(currentTable.item(selectedRow, 1).text())
        longitude = float(currentTable.item(selectedRow, 2).text())

        # Mostra le coordinate del punto selezionato nella status bar
        self.statusBar().showMessage(
            f"Coordinate: {latitude}, {longitude}")

    def show_on_map(self):
        # Recupera la tabella selezionata e la riga selezionata
        currentTab = self.mapTabs.currentIndex()
        currentTable = self.mapTabs.currentWidget().widget(currentTab)
        selectedRows = currentTable.selectionModel().selectedRows()
        if len(selectedRows) == 0:
            return

        # Recupera le coordinate del punto selezionato
        selectedRow = selectedRows[0].row()
        latitude = float(currentTable.item(selectedRow, 1).text())
        longitude = float(currentTable.item(selectedRow, 2).text())

        # Creazione del widget per la mappa
        mapWidget = QWebEngineView()
        self.map_path2 = os.path.abspath("map2.html")
        #self.map_html.save(self.map_path2)
        mapWidget.load(QUrl(self.map_path2))

        # Caricamento della pagina HTML nella QWebEngineView
        html = f'''
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8" />
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>My Map</title>
                
                <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/leaflet@1.9.3/dist/leaflet.css"/>
                <script src="https://cdn.jsdelivr.net/npm/leaflet@1.9.3/dist/leaflet.js"></script>
                
                <style>
                    #map {{ 
                        height: 100vh; 
                        width: 100%; 
                    }}
                </style>
            
            </head>
            <body>
                <div id="map"></div>
                <script>
                    var map = L.map('map').setView([{latitude},{longitude}], 13);
                    L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
                        attribution: 'Map data © <a href="https://openstreetmap.org">OpenStreetMap</a> contributors',
                        maxZoom: 18,
                    }}).addTo(map);
                    // Aggiungi un marker alle coordinate selezionate
                    L.marker([{latitude},{longitude}]).addTo(map);
                </script>
            </body>
            </html>
            '''


        mapWidget.setHtml(html, QUrl('file://'))
        print(html)
        # Aggiunge il nuovo widget alla tab
        selectedName = currentTable.item(selectedRow, 0).text()
        self.mapTabs.addTab(mapWidget, f"{selectedName} sulla Mappa")

        # Seleziona la nuova tab
        newTab = self.mapTabs.count() - 1
        self.mapTabs.setCurrentIndex(newTab)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    mainWindow.show()
    sys.exit(app.exec_())
